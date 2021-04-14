'''

Excelのパラメータ一覧を読み込み、jinja2テンプレートに適用して出力する
Excelの1行目がJinja2の変数名を表し、2行目以降を1行ずつ処理する。

'''

import jinja2 as j2
from pathlib import Path
import openpyxl as px
import tkinter.filedialog as fd
from tkinter import messagebox
import tkinter as tk
import argparse
import sys
import datetime as dt


if __name__ == '__main__':
    tk.Tk().withdraw()
    # コマンドラインプロンプト設定
    parser = argparse.ArgumentParser(
        description=' 機能：Excelのパラメータ一覧を読み込み、jinja2テンプレートに適用してファイルを出力する'
    )
    parser.add_argument(
        'parameter', metavar='filename.xlsx', type=str, nargs='?', help='パラメータExcelファイル'
    )
    parser.add_argument(
        'template', metavar='filename.j2', type=str, nargs='?', help='jinja2テンプレートファイル'
    )
    args = parser.parse_args()

    # パラメータファイル名の取得
    if args.parameter:
        parameter_filename = args.parameter
    else:
        parameter_filename = fd.askopenfilename(
            title='パラメータExcelファイルを選択',
            filetypes=[('xlsxファイル', '*.xlsx'), ('すべて', '*.*')],
            initialdir='.',
        )
        if parameter_filename == '':
            sys.exit(0)

    # テンプレートファイル名の取得
    if args.template:
        template_filename = args.template
    else:
        template_filename = fd.askopenfilename(
            title='jinja2テンプレートファイルを選択',
            filetypes=[('jinja2テンプレート', '*.j2;*.template'), ('すべて', '*.*')],
            initialdir='.',
        )
        if template_filename == '':
            sys.exit(0)

    # Jinja2テンプレート読み込み
    template_filename = Path(template_filename)
    env = j2.Environment(loader=j2.FileSystemLoader(str(template_filename.parent)))
    try:
        template = env.get_template(template_filename.name)
    except j2.TemplateSyntaxError as e:
        message = f'{e.message}\n'
        message += f' File "{e.filename}", line {e.lineno}, in template\n'
        message += f'  {e.source.splitlines()[e.lineno-1]}'
        messagebox.showerror('TemplateSyntaxError', message)
        sys.exit(0)
    except j2.TemplateNotFound as e:
        message = f'Template "{e.name}" not found\n'
        messagebox.showerror('TemplateNotFound', message)
        sys.exit(0)

    # パラメータの読み込み＆レンダリング
    try:
        wb = px.load_workbook(parameter_filename, data_only=True)
        ws = wb.worksheets[0]   # 先頭シートのみを対象とする
        header = ws[1]  # 先頭行はヘッダ、jinja2の変数名になる
    except Exception as e:
        messagebox.showerror('OpenpyxlError', e)
        sys.exit(0)

    # 出力ファイル名に付加する日時文字列
    datetime = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = Path(f'output_{datetime}')
    output_dir.mkdir()

    def create_filename(parent_dir, filename, ext='txt'):
        output_file = parent_dir / Path(f'{filename}.{ext}')
        # ファイルが存在している場合、ファイル名に_\dを追加する。
        counter = 1
        while output_file.exists():
            counter += 1
            output_file = parent_dir / Path(f'{filename}_{counter}.{ext}')
        return output_file

    # Excelを1行ずつ処理する
    for row in ws.iter_rows(min_row=2):
        parameters = {str(name.value): data.value for name, data in zip(header, row)}
        parameters['template_filename'] = template_filename.name
        # parametersに'filename'があれば、それを出力ファイル名とする。無ければ'output'とする。
        output_file = create_filename(output_dir, parameters.get('filename', 'output'))
        with output_file.open('wt') as f:
            print(f'output: {output_file}')
            print(template.render(parameters), file=f)

    messagebox.showinfo(Path(sys.argv[0]).name, f'完了\n{ws.max_row - 1}ファイル出力')
